from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
from tkcalendar import Calendar
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import re
from qr_code_generator import generate_qr_code  # Import the QR code generation function

# Create the workbook if it doesn't exist
file = pathlib.Path('User_record.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    sheet['F1'] = "Email"
    sheet['G1'] = "Date of Birth"
    sheet['H1'] = "Website URL"
    file.save('User_record.xlsx')


def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)
    email = emailValue.get()
    dob = dobEntry.get()  # Get the value from the Date of Birth field
    url = urlValue.get()  # Get the URL

    # Validation
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showerror("Error", "Invalid email format")
        return
    if not contact.isdigit() or len(contact) != 10:
        messagebox.showerror("Error", "Invalid phone number. Enter 10 digits.")
        return
    if url and not re.match(r"http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+", url):
        messagebox.showerror("Error", "Invalid URL format.")
        return

    try:
        # Collect all the data as a single string
        data = f"Name: {name}\nPhone: {contact}\nAge: {age}\nGender: {gender}\nAddress: {address}\nEmail: {email}\nDate of Birth: {dob}\nWebsite: {url}"

        # Save the data in Excel
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        sheet.cell(column=1, row=sheet.max_row+1, value=name)
        sheet.cell(column=2, row=sheet.max_row, value=contact)
        sheet.cell(column=3, row=sheet.max_row, value=age)
        sheet.cell(column=4, row=sheet.max_row, value=gender)
        sheet.cell(column=5, row=sheet.max_row, value=address)
        sheet.cell(column=6, row=sheet.max_row, value=email)
        sheet.cell(column=7, row=sheet.max_row, value=dob)
        sheet.cell(column=8, row=sheet.max_row, value=url)
        workbook.save(file)

        # Generate and save the QR code
        generate_qr_code(data, contact)  # Call the QR code generation function

        messagebox.showinfo("Success", "Data submitted successfully and QR code generated!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    emailValue.set('')
    urlValue.set('')
    addressEntry.delete(1.0, END)
    dobEntry.delete(0, END)


def open_calendar():
    top = Toplevel(root)
    top.title("Select Date")

    # Create the calendar widget
    cal = Calendar(top, selectmode="day", year=2000, month=1, day=1)
    cal.pack(pady=20)

    def grab_date():
        dobEntry.delete(0, END)
        dobEntry.insert(0, cal.get_date())  # Get the selected date and insert into the entry
        top.destroy()  # Close the calendar window

    Button(top, text="Select", command=grab_date).pack(pady=20)


root = tk.Tk()
root.title("Data Entry")
root.geometry('800x600+300+200')
root.resizable(False, False)
root.configure(bg="#87CEEB")

# Icon
icon_image = PhotoImage(file="icon.png")
root.iconphoto(False, icon_image)

# Heading
Label(root, text="Please fill This form: ", font="arial", bg="#87CEEB", fg="#000000").place(x=20, y=20)

# Labels
Label(root, text='Name: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=80)
Label(root, text='Contact: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=120)
Label(root, text='Age:', font=23, bg="#87CEEB", fg="black").place(x=20, y=160)
Label(root, text='Gender: ', font=23, bg="#87CEEB", fg="black").place(x=315, y=160)
Label(root, text='Address: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=200)
Label(root, text='Email: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=260)
Label(root, text='Date of Birth: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=300)
Label(root, text='Website URL: ', font=23, bg="#87CEEB", fg="black").place(x=20, y=340)  # New URL label

# Entry fields
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()
emailValue = StringVar()
urlValue = StringVar()  # New URL variable

nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=15, bd=2, font=20)
ageEntry = Entry(root, textvariable=AgeValue, width=15, bd=2, font=20)
emailEntry = Entry(root, textvariable=emailValue, width=37, bd=2, font=20)
urlEntry = Entry(root, textvariable=urlValue, width=37, bd=2, font=20)  # New URL entry
addressEntry = Text(root, width=51, height=2, bd=2)
dobEntry = Entry(root, width=15, bd=2, font=20)  # Text field for date of birth

nameEntry.place(x=100, y=80)
contactEntry.place(x=100, y=120)
ageEntry.place(x=100, y=160)
addressEntry.place(x=100, y=200)
emailEntry.place(x=100, y=260)
dobEntry.place(x=150, y=300)
urlEntry.place(x=150, y=340)  # Place the URL entry

# Gender
gender_combobox = Combobox(root, values=[' Male', ' Female', ' Other'], font='arial 13', state='r', width=10)
gender_combobox.place(x=400, y=160)
gender_combobox.set(' Male')

# Bind date of birth field to open the calendar when clicked
dobEntry.bind("<1>", lambda e: open_calendar())

# Submit and Clear buttons
Button(root, text="Submit", bg="#87CEEB", fg="black", width=15, height=2, command=submit).place(x=100, y=450)
Button(root, text="Clear", bg="#87CEEB", fg="black", width=15, height=2, command=clear).place(x=250, y=450)
Button(root, text="Exit", bg="#87CEEB", fg="black", width=15, height=2, command=lambda: root.destroy()).place(x=400, y=450)

root.mainloop()
