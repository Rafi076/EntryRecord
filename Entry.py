from tkinter import *
from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

file = pathlib.Path('User_record.xlsx')
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    
    file.save('User_record.xlsx')


def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)
    
    try:
        workbook = openpyxl.load_workbook(file)  # Corrected
        sheet = workbook.active
        sheet.cell(column=1, row=sheet.max_row+1, value=name)
        sheet.cell(column=2, row=sheet.max_row, value=contact)
        sheet.cell(column=3, row=sheet.max_row, value=age)
        sheet.cell(column=4, row=sheet.max_row, value=gender)
        sheet.cell(column=5, row=sheet.max_row, value=address)
        workbook.save(file)
        messagebox.showinfo("Success", "Data submitted successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")


def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0,END)

root = tk.Tk()  # Corrected
root.title("Data Entry")
root.geometry('700x400+300+200')
root.resizable(False, False)  # Corrected spelling of resizable
root.configure(bg="#87CEEB")

#icon
icon_image = PhotoImage(file="icon.png")
root.iconphoto(False,icon_image)

# Heading
Label(root,text="Please fill This form: ",font="arial",bg="#87CEEB",fg="#000000").place(x=20,y=20)

# label
Label(root,text='Name:',font=23,bg="#87CEEB",fg="black").place(x=20,y=80)
Label(root,text='Contact:',font=23,bg="#87CEEB",fg="black").place(x=20,y=120)
Label(root,text='Age:',font=23,bg="#87CEEB",fg="black").place(x=20,y=160)
Label(root,text='Gender:',font=23,bg="#87CEEB",fg="black").place(x=340,y=160)
Label(root,text='Address:',font=23,bg="#87CEEB",fg="black").place(x=20,y=200)

# Entry
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()


nameEntry = Entry(root,textvariable=nameValue, width=45, bd=2, font=20)
contactEntry = Entry(root,textvariable=contactValue, width=45, bd=2, font=20)
ageEntry = Entry(root,textvariable=AgeValue, width=10, bd=2, font=20)
addressEntry = Text(root,width=51,height=3,bd=2)

nameEntry.place(x=100, y=80)
contactEntry.place(x=100, y=120)
ageEntry.place(x=100, y=160)
addressEntry.place(x=100,y=200)

## Gender
gender_combobox = Combobox(root,values=['Male','Female','Other'],font='arial 13',state='r',width=10)
gender_combobox.place(x=400,y=160)
gender_combobox.set('Male') # Male seted as default

#submit button
Button(root,text="Submit",bg="#87CEEB",fg="black",width=15,height=2,command=submit).place(x=100,y=300)
Button(root,text="Clear",bg="#87CEEB",fg="black",width=15,height=2,command=clear).place(x=250,y=300)
Button(root,text="Exit",bg="#87CEEB",fg="black",width=15,height=2,command=lambda:root.destroy()).place(x=400,y=300)

root.mainloop()